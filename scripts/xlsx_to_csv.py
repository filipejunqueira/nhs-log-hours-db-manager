"""xlsx_to_csv.py -- the workbook is the source; the engine still eats CSV.

    python3 scripts/xlsx_to_csv.py <workbook.xlsx> <out-dir>

Writes <out-dir>/hours.csv and, if the workbook has a payments tab,
<out-dir>/payments.csv, in exactly the shape the engine already accepts. The
spreadsheet exports one tab per CSV but downloads whole, so taking the workbook
is the only way to get both tabs in one action.

THIS COMPUTES NOTHING. It reformats what the spreadsheet holds. Minutes are
copied across, never recalculated from Start and End -- the engine cross-checks
those two against each other, and that check only means something if the two
numbers arrive independently.

The output is byte-comparable with the CSVs the pipeline already archives:
CRLF line endings, dates as 1-Jun-26, times as 08:45, Hours to two decimals,
and the header row copied verbatim from the sheet, leading space and all.
"""

import csv
import sys
from datetime import date, datetime, time

import openpyxl

# A sheet is identified by what its header says, not by its name -- the same way
# ingest.sh has always recognised a CSV by its header rather than its filename.
HOURS_COLS = ("date", "start", "end", "minutes", "hours")
PAY_COLS = ("date", "minutespaid")


def _norm(v) -> str:
    return (
        str(v).strip().lower().replace(" ", "").replace("_", "")
        if v is not None
        else ""
    )


def classify(header: list) -> str | None:
    cols = tuple(_norm(c) for c in header)
    if cols[: len(HOURS_COLS)] == HOURS_COLS:
        return "hours"
    if cols[: len(PAY_COLS)] == PAY_COLS:
        return "payments"
    return None


def die(sheet: str, rownum: int, msg: str):
    raise SystemExit(f"ERROR: sheet '{sheet}' row {rownum}: {msg}")


def fmt_date(v, sheet, rownum, col):
    if isinstance(v, datetime):
        v = v.date()
    if not isinstance(v, date):
        die(sheet, rownum, f"{col} is {v!r}, which is not a date")
    return f"{v.day}-{v:%b}-{v:%y}"


def fmt_time(v, sheet, rownum, col):
    if isinstance(v, datetime):
        v = v.time()
    if isinstance(v, time):
        return f"{v.hour:02d}:{v.minute:02d}"
    # A cell that lost its time formatting comes back as the underlying
    # fraction of a day (8:45 is stored as 0.364583...). Accept it rather than
    # refuse a file that is really fine.
    #
    # Hours are zero-padded (08:45). The canonical CSV is INCONSISTENT here --
    # 32 of its 47 rows pad the Start hour and 15 do not -- so no single rule
    # reproduces it byte for byte. Proven 2026-08-18 that the engine computes
    # every day identically either way, so this is cosmetic; padding is chosen
    # because it matches the larger group, making the one-time reformat 15 rows
    # instead of 32, and because from here on the converter is the only writer,
    # so the file stays consistent.
    if isinstance(v, (int, float)):
        mins = round(float(v) * 1440)
        return f"{mins // 60:02d}:{mins % 60:02d}"
    die(sheet, rownum, f"{col} is {v!r}, which is not a time")


def fmt_int(v, sheet, rownum, col):
    # Minutes and Hours are FORMULA cells in the real workbook. openpyxl with
    # data_only=True hands back the value the spreadsheet last calculated, so a
    # workbook written by something that does not evaluate formulas arrives with
    # these empty. That must stop the run, never pass a blank through.
    if v is None or v == "":
        die(sheet, rownum, f"{col} is empty (a formula that was never calculated?)")
    try:
        return str(int(round(float(v))))
    except (TypeError, ValueError):
        die(sheet, rownum, f"{col} is {v!r}, which is not a number")


def fmt_hours(v, sheet, rownum, col):
    if v is None or v == "":
        die(sheet, rownum, f"{col} is empty (a formula that was never calculated?)")
    try:
        return f"{float(v):.2f}"
    except (TypeError, ValueError):
        die(sheet, rownum, f"{col} is {v!r}, which is not a number")


def text(v) -> str:
    return "" if v is None else str(v)


def convert_hours(ws, out_path: str) -> tuple:
    rows = ws.iter_rows()
    header = [c.value for c in next(rows)]
    written, dates = 0, []
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)  # default dialect: CRLF, quote only when needed
        w.writerow([text(h) for h in header])
        for r in rows:
            cells = list(r) + [None] * (6 - len(r))
            # Blank Date means a trailing empty row, which spreadsheets are full
            # of. Skip it rather than write an empty CSV line.
            if cells[0].value is None:
                continue
            n = cells[0].row
            d = fmt_date(cells[0].value, ws.title, n, "Date")
            w.writerow(
                [
                    d,
                    fmt_time(cells[1].value, ws.title, n, "Start"),
                    fmt_time(cells[2].value, ws.title, n, "End"),
                    fmt_int(cells[3].value, ws.title, n, "Minutes"),
                    fmt_hours(cells[4].value, ws.title, n, "Hours"),
                    text(cells[5].value if len(cells) > 5 else None),
                ]
            )
            written += 1
            dates.append(d)
    return written, dates


def convert_payments(ws, out_path: str) -> tuple:
    rows = ws.iter_rows()
    header = [c.value for c in next(rows)]
    written, dates = 0, []
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([text(h) for h in header])
        for r in rows:
            cells = list(r) + [None] * (4 - len(r))
            vals = [c.value if c is not None else None for c in cells[:4]]
            if vals[0] is None and vals[1] is None:
                continue
            n = cells[0].row
            if vals[0] is None or vals[1] is None:
                die(ws.title, n, "a payment needs both Date and MinutesPaid")
            d = fmt_date(vals[0], ws.title, n, "Date")
            w.writerow(
                [
                    d,
                    fmt_int(vals[1], ws.title, n, "MinutesPaid"),
                    fmt_hours(vals[2], ws.title, n, "HoursPaid")
                    if vals[2] is not None
                    else "",
                    text(vals[3]),
                ]
            )
            written += 1
            dates.append(d)
    return written, dates


def sheets_in(book: str) -> dict:
    """{kind: worksheet} for whichever of the two known tabs the workbook has."""
    wb = openpyxl.load_workbook(book, data_only=True, read_only=False)
    found = {}
    for name in wb.sheetnames:
        ws = wb[name]
        first = next(ws.iter_rows(max_row=1), None)
        kind = classify([c.value for c in first]) if first else None
        if kind and kind not in found:
            found[kind] = ws
    return found


def probe(book: str) -> int:
    """--probe: is this an hours workbook? Lets ingest.sh recognise a file by
    content without a second copy of the header rule living in the shell."""
    try:
        found = sheets_in(book)
    except Exception:
        return 1
    if "hours" not in found:
        return 1
    print(" ".join(sorted(found)))
    return 0


def main() -> None:
    if len(sys.argv) == 3 and sys.argv[1] == "--probe":
        raise SystemExit(probe(sys.argv[2]))
    if len(sys.argv) != 3:
        raise SystemExit(
            f"usage: {sys.argv[0]} <workbook.xlsx> <out-dir>\n"
            f"       {sys.argv[0]} --probe <workbook.xlsx>"
        )
    book, out_dir = sys.argv[1], sys.argv[2]

    # data_only=True (inside sheets_in) gives the last calculated value of a
    # formula cell rather than the formula text. See fmt_int for what happens
    # when there isn't one.
    found = sheets_in(book)
    if "hours" not in found:
        raise SystemExit(
            f"ERROR: no sheet in {book} has an hours header "
            f"(Date, Start, End, Minutes, Hours, ...)"
        )

    n, dates = convert_hours(found["hours"], f"{out_dir}/hours.csv")
    print(
        f"hours    : {n} rows -> {out_dir}/hours.csv"
        + (f" (to {dates[-1]})" if dates else "")
    )

    if "payments" in found:
        n, dates = convert_payments(found["payments"], f"{out_dir}/payments.csv")
        print(
            f"payments : {n} rows -> {out_dir}/payments.csv"
            + (f" (to {dates[-1]})" if dates else " (none recorded yet)")
        )
    else:
        print(f"payments : no payments sheet in {book}; nothing written")


if __name__ == "__main__":
    main()
